"""
This will be a simulation for a general scenario in video games to compare two sets of stats and calculate their theoretical or average Damage (DMG).
The Damage for a single hit (DMG) formula is as follows:
DMG = Attack (ATK) * Bonuses
Bonuses are calculated as:
Bonuses = %DMG Bonus * Critical Damage
%DMG Bonus is calculated as:
%DMG Bonus = 1 + (Elemental Bonus + Skill Bonus + Other Bonuses)
Critical Damage is calculated as:
Critical Damage = 1 + (CC * CD)
Where CC is the Critical Hit Chance (in percent) and CD is the Critical Hit Damage (in percent).
This code will prompt the user to input two iterations of values for ATK (Flat Value), Elemental Bonus (in percent), Skill Bonus (in percent), Other Bonuses (in percent), CC (in percent), and CD (in percent).
It will then calculate the Damage for each set and compare them, outputting the results in a table format.
"""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Function to calculate expected damage (average over many hits)
def calculate_expected_damage(atk, elemental_bonus, skill_bonus, other_bonuses, crit_chance, crit_damage):
    percent_dmg_bonus = 1 + (elemental_bonus + skill_bonus + other_bonuses) / 100
    crit_multiplier = 1 + (crit_chance / 100) * (crit_damage / 100)
    damage = atk * percent_dmg_bonus * crit_multiplier
    return damage

# Function to calculate damage for a critical hit
def calculate_crit_damage(atk, elemental_bonus, skill_bonus, other_bonuses, crit_damage):
    percent_dmg_bonus = 1 + (elemental_bonus + skill_bonus + other_bonuses) / 100
    crit_multiplier = 1 + (crit_damage / 100)
    damage = atk * percent_dmg_bonus * crit_multiplier
    return damage

# Function to calculate damage for a normal hit (non-crit)
def calculate_normal_damage(atk, elemental_bonus, skill_bonus, other_bonuses):
    percent_dmg_bonus = 1 + (elemental_bonus + skill_bonus + other_bonuses) / 100
    damage = atk * percent_dmg_bonus
    return damage

# Function to get user input for stats
def get_stats_input(iteration):
    print(f"\nEnter stats for iteration {iteration}:")
    atk = float(input("Attack (ATK) flat value: "))
    elemental_bonus = float(input("Elemental Bonus (%): "))
    skill_bonus = float(input("Skill Bonus (%): "))
    other_bonuses = float(input("Other Bonuses (%): "))
    crit_chance = float(input("Criticaltesr(CC) (%): "))
    crit_damage = float(input("Critical Hit Damage (CD) (%): "))
    return atk, elemental_bonus, skill_bonus, other_bonuses, crit_chance, crit_damage

# Get stats for two iterations
stats1 = get_stats_input(1)
stats2 = get_stats_input(2)

# Calculate damage for both sets of stats
expected_damage1 = calculate_expected_damage(*stats1)
expected_damage2 = calculate_expected_damage(*stats2)

crit_damage1 = calculate_crit_damage(stats1[0], stats1[1], stats1[2], stats1[3], stats1[5])
crit_damage2 = calculate_crit_damage(stats2[0], stats2[1], stats2[2], stats2[3], stats2[5])

normal_damage1 = calculate_normal_damage(stats1[0], stats1[1], stats1[2], stats1[3])
normal_damage2 = calculate_normal_damage(stats2[0], stats2[1], stats2[2], stats2[3])

# Prepare results for comparison
results = {
    "Stat": ["Attack (ATK)", "Elemental Bonus (%)", "Skill Bonus (%)", "Other Bonuses (%)", "Critical Hit Chance (CC) (%)", "Critical Hit Damage (CD) (%)", "Normal Hit Damage", "Critical Hit Damage", "Expected Average Damage"],
    "Iteration 1": [f"{stats1[0]:.2f}", f"{stats1[1]:.2f}", f"{stats1[2]:.2f}", f"{stats1[3]:.2f}", f"{stats1[4]:.2f}", f"{stats1[5]:.2f}", f"{normal_damage1:.2f}", f"{crit_damage1:.2f}", f"{expected_damage1:.2f}"],
    "Iteration 2": [f"{stats2[0]:.2f}", f"{stats2[1]:.2f}", f"{stats2[2]:.2f}", f"{stats2[3]:.2f}", f"{stats2[4]:.2f}", f"{stats2[5]:.2f}", f"{normal_damage2:.2f}", f"{crit_damage2:.2f}", f"{expected_damage2:.2f}"],
}

# Create a DataFrame for better visualization
df_results = pd.DataFrame(results)
print("\nComparison of Damage Calculations:")
print(df_results.to_string(index=False))

# Determine which iteration has higher expected damage
if expected_damage1 > expected_damage2:
    print("\nIteration 1 has higher expected average damage.")
elif expected_damage2 > expected_damage1:
    print("\nIteration 2 has higher expected average damage.")
else:
    print("\nBoth iterations have equal expected average damage.")

# Create Excel workbook with formulas
wb = Workbook()
ws = wb.active
ws.title = "Damage Calculator"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
title_font = Font(bold=True, color="FFFFFF", size=12)
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(bold=True, size=11)

# Title
ws["A1"] = "DAMAGE CALCULATION FORMULAS"
ws["A1"].fill = title_fill
ws["A1"].font = title_font
ws.merge_cells("A1:C1")
ws["A1"].alignment = Alignment(horizontal="center")

row = 3

# Formulas section
ws[f"A{row}"] = "Normal Hit Damage Formula:"
ws[f"A{row}"].font = section_font
ws[f"A{row}"].fill = section_fill
row += 1

ws[f"A{row}"] = "%DMG Bonus = 1 + (Elemental Bonus + Skill Bonus + Other Bonuses) / 100"
row += 1

ws[f"A{row}"] = "Normal DMG = ATK * %DMG Bonus"
row += 2

ws[f"A{row}"] = "Critical Hit Damage Formula:"
ws[f"A{row}"].font = section_font
ws[f"A{row}"].fill = section_fill
row += 1

ws[f"A{row}"] = "Crit Multiplier = 1 + (CD / 100)"
row += 1

ws[f"A{row}"] = "Crit DMG = ATK * %DMG Bonus * Crit Multiplier"
row += 2

ws[f"A{row}"] = "Expected Average Damage Formula:"
ws[f"A{row}"].font = section_font
ws[f"A{row}"].fill = section_fill
row += 1

ws[f"A{row}"] = "Expected Crit Multiplier = 1 + (CC / 100) * (CD / 100)"
row += 1

ws[f"A{row}"] = "Expected Avg DMG = ATK * %DMG Bonus * Expected Crit Multiplier"
row += 3

# Input section header
ws[f"A{row}"] = "INPUT VALUES"
ws[f"A{row}"].fill = title_fill
ws[f"A{row}"].font = title_font
ws.merge_cells(f"A{row}:C{row}")
ws[f"A{row}"].alignment = Alignment(horizontal="center")
row += 1

# Column headers
ws[f"A{row}"] = "Stat"
ws[f"B{row}"] = "Iteration 1"
ws[f"C{row}"] = "Iteration 2"
for col in ["A", "B", "C"]:
    ws[f"{col}{row}"].fill = header_fill
    ws[f"{col}{row}"].font = header_font
row += 1

# Input values
input_row_start = row
ws[f"A{row}"] = "Attack (ATK)"
ws[f"B{row}"] = stats1[0]
ws[f"C{row}"] = stats2[0]
row += 1

ws[f"A{row}"] = "Elemental Bonus (%)"
ws[f"B{row}"] = stats1[1]
ws[f"C{row}"] = stats2[1]
row += 1

ws[f"A{row}"] = "Skill Bonus (%)"
ws[f"B{row}"] = stats1[2]
ws[f"C{row}"] = stats2[2]
row += 1

ws[f"A{row}"] = "Other Bonuses (%)"
ws[f"B{row}"] = stats1[3]
ws[f"C{row}"] = stats2[3]
row += 1

ws[f"A{row}"] = "Critical Hit Chance (CC) (%)"
ws[f"B{row}"] = stats1[4]
ws[f"C{row}"] = stats2[4]
row += 1

ws[f"A{row}"] = "Critical Hit Damage (CD) (%)"
ws[f"B{row}"] = stats1[5]
ws[f"C{row}"] = stats2[5]
row += 3

# Calculated results header
ws[f"A{row}"] = "CALCULATED RESULTS"
ws[f"A{row}"].fill = title_fill
ws[f"A{row}"].font = title_font
ws.merge_cells(f"A{row}:C{row}")
ws[f"A{row}"].alignment = Alignment(horizontal="center")
row += 1

# Column headers for results
ws[f"A{row}"] = "Stat"
ws[f"B{row}"] = "Iteration 1"
ws[f"C{row}"] = "Iteration 2"
for col in ["A", "B", "C"]:
    ws[f"{col}{row}"].fill = header_fill
    ws[f"{col}{row}"].font = header_font
row += 1

# Formulas for calculated results
atk1_cell = f"B{input_row_start}"
elem1_cell = f"B{input_row_start+1}"
skill1_cell = f"B{input_row_start+2}"
other1_cell = f"B{input_row_start+3}"
cc1_cell = f"B{input_row_start+4}"
cd1_cell = f"B{input_row_start+5}"

atk2_cell = f"C{input_row_start}"
elem2_cell = f"C{input_row_start+1}"
skill2_cell = f"C{input_row_start+2}"
other2_cell = f"C{input_row_start+3}"
cc2_cell = f"C{input_row_start+4}"
cd2_cell = f"C{input_row_start+5}"

# Normal Hit Damage
ws[f"A{row}"] = "Normal Hit Damage"
ws[f"B{row}"] = f"={atk1_cell}*(1+({elem1_cell}+{skill1_cell}+{other1_cell})/100)"
ws[f"C{row}"] = f"={atk2_cell}*(1+({elem2_cell}+{skill2_cell}+{other2_cell})/100)"
row += 1

# Critical Hit Damage
ws[f"A{row}"] = "Critical Hit Damage"
ws[f"B{row}"] = f"={atk1_cell}*(1+({elem1_cell}+{skill1_cell}+{other1_cell})/100)*(1+{cd1_cell}/100)"
ws[f"C{row}"] = f"={atk2_cell}*(1+({elem2_cell}+{skill2_cell}+{other2_cell})/100)*(1+{cd2_cell}/100)"
row += 1

# Expected Average Damage
ws[f"A{row}"] = "Expected Average Damage"
ws[f"B{row}"] = f"={atk1_cell}*(1+({elem1_cell}+{skill1_cell}+{other1_cell})/100)*(1+({cc1_cell}/100)*({cd1_cell}/100))"
ws[f"C{row}"] = f"={atk2_cell}*(1+({elem2_cell}+{skill2_cell}+{other2_cell})/100)*(1+({cc2_cell}/100)*({cd2_cell}/100))"

# Set column widths
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 20

# Prompt for custom filename
file_format = input("\nChoose file format (1=Excel, 2=CSV): ").strip()
file_format = file_format if file_format in ["1", "2"] else "1"

filename = input("Enter the name for the file (without extension): ").strip()
if not filename:
    filename = "damage_comparison_results"

if file_format == "1":
    # Save as Excel
    filename = f"{filename}.xlsx"
    wb.save(filename)
    print(f"\nResults have been saved to '{filename}'.")
    print("You can now edit the input values in the 'INPUT VALUES' section and the calculations will update automatically!")
else:
    # Save as CSV
    filename = f"{filename}.csv"
    csv_rows = [
        ["DAMAGE CALCULATION FORMULAS", "", ""],
        ["", "", ""],
        ["Normal Hit Damage Formula:", "", ""],
        ["%DMG Bonus = 1 + (Elemental Bonus + Skill Bonus + Other Bonuses) / 100", "", ""],
        ["Normal DMG = ATK * %DMG Bonus", "", ""],
        ["", "", ""],
        ["Critical Hit Damage Formula:", "", ""],
        ["Crit Multiplier = 1 + (CD / 100)", "", ""],
        ["Crit DMG = ATK * %DMG Bonus * Crit Multiplier", "", ""],
        ["", "", ""],
        ["Expected Average Damage Formula:", "", ""],
        ["Expected Crit Multiplier = 1 + (CC / 100) * (CD / 100)", "", ""],
        ["Expected Avg DMG = ATK * %DMG Bonus * Expected Crit Multiplier", "", ""],
        ["", "", ""],
        ["INPUT VALUES", "", ""],
        ["Stat", "Iteration 1", "Iteration 2"],
        ["Attack (ATK)", f"{stats1[0]:.2f}", f"{stats2[0]:.2f}"],
        ["Elemental Bonus (%)", f"{stats1[1]:.2f}", f"{stats2[1]:.2f}"],
        ["Skill Bonus (%)", f"{stats1[2]:.2f}", f"{stats2[2]:.2f}"],
        ["Other Bonuses (%)", f"{stats1[3]:.2f}", f"{stats2[3]:.2f}"],
        ["Critical Hit Chance (CC) (%)", f"{stats1[4]:.2f}", f"{stats2[4]:.2f}"],
        ["Critical Hit Damage (CD) (%)", f"{stats1[5]:.2f}", f"{stats2[5]:.2f}"],
        ["", "", ""],
        ["CALCULATED RESULTS", "", ""],
        ["Stat", "Iteration 1", "Iteration 2"],
        ["Normal Hit Damage", f"{normal_damage1:.2f}", f"{normal_damage2:.2f}"],
        ["Critical Hit Damage", f"{crit_damage1:.2f}", f"{crit_damage2:.2f}"],
        ["Expected Average Damage", f"{expected_damage1:.2f}", f"{expected_damage2:.2f}"],
    ]
    df_csv = pd.DataFrame(csv_rows, columns=["Description", "Iteration 1", "Iteration 2"])
    df_csv.to_csv(filename, index=False)
    print(f"\nResults have been saved to '{filename}'.")

# End of DPScompare.py